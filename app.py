from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def clean_header(col):
    return col.replace("\n", " ").strip()


def safe_key(col):
    return clean_header(col).lower().replace(" ", "_")


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["file"]

        df = pd.read_excel(file).fillna("")
        df.columns = [clean_header(c) for c in df.columns]

        column_map = {c: safe_key(c) for c in df.columns}

        return render_template(
            "table.html",
            columns=df.columns.tolist(),
            column_map=column_map,
            data=df.to_dict(orient="records")
        )

    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():
    columns = request.form.getlist("columns")
    safe_columns = request.form.getlist("safe_columns")
    row_count = int(request.form["row_count"])
    action = request.form.get("action")

    kitchen_name = request.form.get("kitchen_name", "Kitchen").replace(" ", "_")
    today = datetime.now().strftime("%Y-%m-%d")

    rows = []

    for i in range(row_count):
        row = {}
        for col, safe in zip(columns, safe_columns):
            val = request.form.get(f"{safe}_{i}", "").strip()
            row[col] = val
        rows.append(row)

    df = pd.DataFrame(rows)

    # Identify important columns
    requested_col = next(c for c in df.columns if "requested" in c.lower())
    supplier_col = next(c for c in df.columns if "supplier" in c.lower())
    current_col = next(c for c in df.columns if "current" in c.lower())

    # Convert requested quantity to numeric
    df[requested_col] = pd.to_numeric(df[requested_col], errors="coerce").fillna(0)

    # Save master inventory (always)
    master_path = os.path.join(
        UPLOAD_FOLDER,
        f"{kitchen_name}_{today}_master_inventory.xlsx"
    )
    df.to_excel(master_path, index=False)

    # Save draft only
    if action == "draft":
        draft_path = os.path.join(
            UPLOAD_FOLDER,
            f"{kitchen_name}_{today}_draft.xlsx"
        )
        df.to_excel(draft_path, index=False)
        return "Draft saved successfully"

    # Filter order items
    order_df = df[df[requested_col] > 0].copy()

    # Remove current quantity from order list
    if current_col in order_df.columns:
        order_df.drop(columns=[current_col], inplace=True)

    # Output path
    output_path = os.path.join(
        UPLOAD_FOLDER,
        f"{kitchen_name}_{today}_order_list.xlsx"
    )

    # Save supplier-wise sheets first
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for supplier, group in order_df.groupby(supplier_col):
            sheet_name = str(supplier)[:31]  # Excel sheet name max 31 chars
            group.to_excel(writer, sheet_name=sheet_name, index=False)

    # Open workbook to apply wrap text and auto column width
    wb = load_workbook(output_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Wrap text
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Auto adjust column width
        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)

    return send_file(output_path, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
